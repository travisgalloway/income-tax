"""Excel parsing for the two spreadsheet sources this pipeline reads.

Mirrors `normalize.parse_cbo`'s contract: an empty parse is a failure, never
"no rows". A workbook with no sheets, or a sheet that parses to zero data rows,
raises SourceUnavailable rather than letting a builder silently chart nothing.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

import openpyxl

from .errors import SourceUnavailable


def sheet_rows(
    content: bytes,
    *,
    source: str,
    url: str,
    sheet: str | None = None,
) -> list[list[Any]]:
    """Return every row of one worksheet as a list of cell values.

    Defaults to the first sheet. `data_only=True` reads cached formula results
    rather than formula text; `read_only=True` streams rather than loading the
    whole workbook into memory.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # a truncated or corrupt .xlsx fails to parse at all
        raise SourceUnavailable(source, url, f"could not parse workbook: {exc}") from exc

    if not wb.sheetnames:
        raise SourceUnavailable(source, url, "workbook has no sheets")

    name = sheet if sheet is not None else wb.sheetnames[0]
    if name not in wb.sheetnames:
        raise SourceUnavailable(
            source, url, f"sheet {name!r} not found; workbook has {wb.sheetnames}"
        )

    ws = wb[name]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    if not rows:
        raise SourceUnavailable(source, url, f"sheet {name!r} parsed to zero rows")
    return rows
